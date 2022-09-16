from bd import Connection
from openpyxl import Workbook

connect = Connection()
cursor = connect.cursor()

# query = '''SELECT p.id, p.primer_apellido FROM padron p'''

query = '''SELECT p.id, p.primer_apellido, p.segundo_apellido, 
           p.nombres, p.comentario FROM padron AS p LIMIT 100'''

cursor.execute(query)
data = cursor.fetchall()

def actualizar(apellido, id):
    cursor = connect.cursor()
    try:
        cursor.execute(f"UPDATE padron SET primer_apellido = '{apellido}' WHERE id = '{id}'") 
    except:
        print('Error en la actualizacion.')    
        connect.rollback()
    connect.commit()
    cursor.close()

# query = 'UPDATE padron SET primer_apellido = "AGUILAR" WHERE id = "1"'
# actualizar(query)

# Hubo un error de espacios en la columna primer_apellido, se realizo la actualizacion
# con el siguiente ciclo

# for i in data:
#     id = i[0]
#     apellido = i[1].strip()
#     actualizar(apellido, id)
#     print(i[0])

wb = Workbook()
ws = wb.active

# titulos = {
#     'A': 'identificador', 
#     'B': 'primer_apellido', 
#     'C': 'segundo_apellido', 
#     'D': 'nombres', 
#     'E': 'comentario'}

# for c, v in titulos.items():
    # print(f'{c}1', v)
    # print(ws['A1'])
    # # celda = 
    # ws[f'{c}1'] = titulos[v]

ws['A1'] = 'Identificador'
ws['B1'] = 'Primer Apellido'
ws['C1'] = 'Segundo Apellido'
ws['D1'] = 'Nombres'
ws['E1'] = 'Comentario'

for i in data:
    ws.append(i)

wb.save('excel_mysql/Primeros cien.xlsx')

